from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Count, F, Q
from django.db import models
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, date
from decimal import Decimal
import calendar


class SalesReportView(APIView):
    """Sales Report API"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.invoices.models import Invoice
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        invoices = Invoice.objects.filter(
            invoice_date__gte=start_date,
            invoice_date__lte=end_date
        ).exclude(status='cancelled')
        
        summary = invoices.aggregate(
            total_invoices=Count('id'),
            total_sales=Sum('total'),
            total_received=Sum('received'),
            total_pending=Sum('balance'),
            total_tax=Sum('tax_amount'),
            total_discount=Sum('discount_amount')
        )
        
        # Daily breakdown
        daily_sales = invoices.values('invoice_date').annotate(
            count=Count('id'),
            total=Sum('total')
        ).order_by('invoice_date')
        
        return Response({
            'period': {'start': start_date, 'end': end_date},
            'summary': {
                'total_invoices': summary['total_invoices'] or 0,
                'total_sales': float(summary['total_sales'] or 0),
                'total_received': float(summary['total_received'] or 0),
                'total_pending': float(summary['total_pending'] or 0),
                'total_tax': float(summary['total_tax'] or 0),
                'total_discount': float(summary['total_discount'] or 0),
            },
            'daily_breakdown': list(daily_sales)
        })


class PurchaseReportView(APIView):
    """Purchase Report API"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.purchases.models import Purchase, Expense
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        purchases = Purchase.objects.filter(
            purchase_date__gte=start_date,
            purchase_date__lte=end_date
        ).exclude(status='cancelled')
        
        purchase_summary = purchases.aggregate(
            total_purchases=Count('id'),
            total_amount=Sum('total'),
            total_paid=Sum('paid'),
            total_pending=Sum('balance')
        )
        
        expenses = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date
        )
        
        expense_summary = expenses.aggregate(
            total_expenses=Count('id'),
            total_amount=Sum('amount')
        )
        
        return Response({
            'period': {'start': start_date, 'end': end_date},
            'purchases': {
                'count': purchase_summary['total_purchases'] or 0,
                'total': float(purchase_summary['total_amount'] or 0),
                'paid': float(purchase_summary['total_paid'] or 0),
                'pending': float(purchase_summary['total_pending'] or 0),
            },
            'expenses': {
                'count': expense_summary['total_expenses'] or 0,
                'total': float(expense_summary['total_amount'] or 0),
            }
        })


class ProfitLossReportView(APIView):
    """Profit & Loss Report API"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.invoices.models import Invoice
        from apps.purchases.models import Purchase, Expense
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        # Income
        sales = Invoice.objects.filter(
            invoice_date__gte=start_date,
            invoice_date__lte=end_date
        ).exclude(status='cancelled').aggregate(
            total=Sum('subtotal')
        )['total'] or 0
        
        # Cost of goods
        purchases = Purchase.objects.filter(
            purchase_date__gte=start_date,
            purchase_date__lte=end_date
        ).exclude(status='cancelled').aggregate(
            total=Sum('subtotal')
        )['total'] or 0
        
        gross_profit = float(sales) - float(purchases)
        
        # Expenses
        expenses = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        net_profit = gross_profit - float(expenses)
        
        # Expense breakdown
        expense_breakdown = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date
        ).values('category').annotate(total=Sum('amount'))
        
        return Response({
            'period': {'start': start_date, 'end': end_date},
            'income': {
                'sales': float(sales),
            },
            'cost_of_goods': {
                'purchases': float(purchases),
            },
            'gross_profit': gross_profit,
            'expenses': {
                'total': float(expenses),
                'breakdown': list(expense_breakdown)
            },
            'net_profit': net_profit,
            'profit_margin': round((net_profit / float(sales) * 100) if sales else 0, 2)
        })


class GSTReportView(APIView):
    """GST Report API"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.invoices.models import Invoice, InvoiceItem
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date:
            start_date = timezone.now().date().replace(day=1)
        if not end_date:
            end_date = timezone.now().date()
        
        invoices = Invoice.objects.filter(
            invoice_date__gte=start_date,
            invoice_date__lte=end_date
        ).exclude(status='cancelled')
        
        # Output tax (sales)
        output_tax = invoices.aggregate(
            cgst=Sum('cgst_amount'),
            sgst=Sum('sgst_amount'),
            igst=Sum('igst_amount'),
            total=Sum('tax_amount')
        )
        
        # Tax rate breakdown
        tax_breakdown = InvoiceItem.objects.filter(
            invoice__invoice_date__gte=start_date,
            invoice__invoice_date__lte=end_date
        ).exclude(invoice__status='cancelled').values('tax_rate').annotate(
            taxable_amount=Sum(F('quantity') * F('price') - F('discount_amount')),
            tax_amount=Sum('tax_amount')
        ).order_by('tax_rate')
        
        return Response({
            'period': {'start': start_date, 'end': end_date},
            'output_tax': {
                'cgst': float(output_tax['cgst'] or 0),
                'sgst': float(output_tax['sgst'] or 0),
                'igst': float(output_tax['igst'] or 0),
                'total': float(output_tax['total'] or 0),
            },
            'rate_breakdown': list(tax_breakdown)
        })


class StockReportView(APIView):
    """Stock/Inventory Report API"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.products.models import Product, Category
        
        try:
            products = Product.objects.filter(is_active=True)
            
            # Summary - handle potential None values
            summary = products.aggregate(
                total_products=Count('id'),
                total_quantity=Sum('quantity'),
                total_value=Sum(F('quantity') * F('purchase_price'))
            )
            
            low_stock = products.filter(quantity__lte=F('min_stock')).count()
            out_of_stock = products.filter(quantity=0).count()
            
            # Category breakdown
            category_breakdown = products.values(
                'category__name'
            ).annotate(
                count=Count('id'),
                quantity=Sum('quantity'),
                value=Sum(F('quantity') * F('purchase_price'))
            )
            
            # Low stock items
            low_stock_items = products.filter(
                quantity__lte=F('min_stock')
            ).values('id', 'name', 'sku', 'quantity', 'min_stock')[:20]
            
            return Response({
                'summary': {
                    'total_products': summary['total_products'] or 0,
                    'total_quantity': summary['total_quantity'] or 0,
                    'total_value': float(summary['total_value'] or 0),
                    'low_stock_count': low_stock,
                    'out_of_stock_count': out_of_stock,
                },
                'category_breakdown': list(category_breakdown),
                'low_stock_items': list(low_stock_items)
            })
        except Exception as e:
            # Return empty data on error to prevent 500
            print(f"StockReportView error: {e}")
            return Response({
                'summary': {
                    'total_products': 0,
                    'total_quantity': 0,
                    'total_value': 0,
                    'low_stock_count': 0,
                    'out_of_stock_count': 0,
                },
                'category_breakdown': [],
                'low_stock_items': []
            })


class DashboardDataView(APIView):
    """Aggregated Dashboard Data API - OPTIMIZED"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.invoices.models import Invoice
        from apps.payments.models import Payment
        from apps.purchases.models import Expense
        from apps.customers.models import Customer
        from apps.products.models import Product
        
        today = timezone.now().date()
        month_start = today.replace(day=1)
        last_month_start = (month_start - timedelta(days=1)).replace(day=1)
        week_ago = today - timedelta(days=6)
        
        # ===== OPTIMIZED: Single query for all invoice aggregations =====
        invoice_stats = Invoice.objects.exclude(status='cancelled').aggregate(
            today_sales=Sum('total', filter=models.Q(invoice_date=today)),
            monthly_sales=Sum('total', filter=models.Q(invoice_date__gte=month_start)),
            last_month_sales=Sum('total', filter=models.Q(
                invoice_date__gte=last_month_start,
                invoice_date__lt=month_start
            )),
            pending_receivable=Sum('balance', filter=models.Q(status__in=['pending', 'partial']))
        )
        
        today_sales = invoice_stats['today_sales'] or 0
        monthly_sales = invoice_stats['monthly_sales'] or 0
        last_month_sales = invoice_stats['last_month_sales'] or 0
        pending_receivable = invoice_stats['pending_receivable'] or 0
        
        # Today's collections
        today_collections = Payment.objects.filter(
            payment_date=today
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # ===== OPTIMIZED: Single query for 7-day sales trend =====
        sales_by_day = Invoice.objects.filter(
            invoice_date__gte=week_ago,
            invoice_date__lte=today
        ).exclude(status='cancelled').values('invoice_date').annotate(
            total=Sum('total')
        ).order_by('invoice_date')
        
        # Build sales_trend with all 7 days (fill missing days with 0)
        sales_dict = {item['invoice_date']: float(item['total']) for item in sales_by_day}
        sales_trend = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            sales_trend.append({
                'date': date.strftime('%Y-%m-%d'),
                'day': date.strftime('%a'),
                'amount': sales_dict.get(date, 0)
            })
        
        # Payment mode breakdown
        payment_modes = Payment.objects.filter(
            payment_date__gte=month_start
        ).values('mode').annotate(total=Sum('amount'))
        
        # Top customers (limited to 5)
        top_customers = Invoice.objects.filter(
            invoice_date__gte=month_start
        ).exclude(status='cancelled').values(
            'customer__id', 'customer__name'
        ).annotate(
            total=Sum('total')
        ).order_by('-total')[:5]
        
        # ===== JOB ORDER METRICS =====
        from apps.joborders.models import JobOrder
        
        total_customers = Customer.objects.count()
        
        # Job order counts by ACTUAL status values (received, finishing, ready)
        job_order_stats = JobOrder.objects.aggregate(
            received=Count('id', filter=models.Q(status='received')),
            finishing=Count('id', filter=models.Q(status='finishing')),
            ready=Count('id', filter=models.Q(status='ready')),
            delivered=Count('id', filter=models.Q(status='delivered')),
            overdue=Count('id', filter=models.Q(
                status__in=['received', 'finishing'],
                expected_delivery__lt=today
            )),
            today_delivery=Count('id', filter=models.Q(
                expected_delivery=today,
                status__in=['received', 'finishing', 'ready']
            ))
        )
        
        # Pending invoice counts for real-time visibility
        pending_invoice_count = Invoice.objects.filter(
            status__in=['pending', 'partial']
        ).exclude(status='cancelled').count()
        
        overdue_invoice_count = Invoice.objects.filter(
            status__in=['pending', 'partial'],
            due_date__lt=today
        ).count()
        
        # Calculate growth safely
        growth = 0
        if last_month_sales:
            growth = round(((float(monthly_sales) - float(last_month_sales)) / float(last_month_sales) * 100), 1)
        
        return Response({
            'today': {
                'sales': float(today_sales),
                'collections': float(today_collections),
                'date': today.strftime('%Y-%m-%d')
            },
            'monthly': {
                'sales': float(monthly_sales),
                'last_month_sales': float(last_month_sales),
                'growth': growth
            },
            'pending': {
                'receivable': float(pending_receivable),
                'invoice_count': pending_invoice_count,
                'overdue_count': overdue_invoice_count,
            },
            'counts': {
                'customers': total_customers,
            },
            'job_orders': {
                'received': job_order_stats['received'] or 0,
                'finishing': job_order_stats['finishing'] or 0,
                'ready': job_order_stats['ready'] or 0,
                'delivered': job_order_stats['delivered'] or 0,
                'overdue': job_order_stats['overdue'] or 0,
                'today_delivery': job_order_stats['today_delivery'] or 0,
            },
            'charts': {
                'sales_trend': sales_trend,
                'payment_modes': list(payment_modes),
                'top_customers': list(top_customers),
            }
        })


def _compute_staff_salary(staff, month, year):
    """Compute salary for a staff member in real-time from attendance data.
    If a saved PaySlip exists, use it. Otherwise, calculate live from Attendance + Holiday records.
    Returns a dict with all salary fields."""
    from apps.staff.models import PaySlip, Attendance, Holiday
    from django.db.models import Sum
    
    payslip = PaySlip.objects.filter(staff=staff, month=month, year=year).first()
    
    if payslip:
        return {
            'id': staff.id,
            'name': staff.name,
            'role': staff.get_role_display(),
            'salary_type': staff.salary_type,
            'daily_rate': float(staff.daily_rate or 0),
            'monthly_salary': float(staff.monthly_salary or 0),
            'payslip_id': payslip.id,
            'total_working_days': payslip.total_working_days,
            'days_present': payslip.days_present,
            'half_days': payslip.half_days,
            'days_absent': payslip.days_absent,
            'leaves': payslip.leaves,
            'sundays': payslip.sundays,
            'holidays_count': payslip.holidays_count,
            'gross_salary': float(payslip.gross_salary),
            'deductions': float(payslip.deductions),
            'overtime_hours': float(payslip.overtime_hours),
            'overtime_amount': float(payslip.overtime_amount),
            'net_salary': float(payslip.net_salary),
            'payment_status': payslip.payment_status,
            'is_live': False,
        }
    
    # ===== Live calculation from Attendance =====
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Count Sundays
    sundays = sum(1 for d in range(1, days_in_month + 1) if date(year, month, d).weekday() == 6)
    
    # Count holidays (exclude those on Sunday)
    holidays_in_month = Holiday.objects.filter(date__month=month, date__year=year)
    holidays_count = sum(1 for h in holidays_in_month if h.date.weekday() != 6)
    
    actual_working_days = days_in_month - sundays - holidays_count
    
    # Attendance counts
    records = Attendance.objects.filter(staff=staff, date__month=month, date__year=year)
    days_present = records.filter(status='present').count()
    half_days = records.filter(status='half_day').count()
    days_absent = records.filter(status='absent').count()
    leaves = records.filter(status='leave').count()
    overtime_hours = records.aggregate(total=Sum('overtime_hours'))['total'] or Decimal('0')
    
    # Calculate salary
    if staff.salary_type == 'daily':
        effective_days = Decimal(str(days_present)) + (Decimal(str(half_days)) * Decimal('0.5'))
        gross_salary = effective_days * (staff.daily_rate or Decimal('0'))
        overtime_rate = (staff.daily_rate or Decimal('0')) / Decimal('8')
        overtime_amount = overtime_hours * overtime_rate
        deductions = Decimal('0')
    else:
        monthly_sal = staff.monthly_salary or Decimal('0')
        if actual_working_days > 0:
            per_day = monthly_sal / Decimal(str(actual_working_days))
        else:
            per_day = Decimal('0')
        
        absent_deduction = Decimal(str(days_absent)) * per_day
        half_day_deduction = Decimal(str(half_days)) * per_day * Decimal('0.5')
        deductions = absent_deduction + half_day_deduction
        gross_salary = monthly_sal
        overtime_rate = per_day / Decimal('8') if per_day > 0 else Decimal('0')
        overtime_amount = overtime_hours * overtime_rate
    
    net_salary = gross_salary - deductions + overtime_amount
    
    return {
        'id': staff.id,
        'name': staff.name,
        'role': staff.get_role_display(),
        'salary_type': staff.salary_type,
        'daily_rate': float(staff.daily_rate or 0),
        'monthly_salary': float(staff.monthly_salary or 0),
        'payslip_id': None,
        'total_working_days': actual_working_days,
        'days_present': days_present,
        'half_days': half_days,
        'days_absent': days_absent,
        'leaves': leaves,
        'sundays': sundays,
        'holidays_count': holidays_count,
        'gross_salary': float(round(gross_salary, 2)),
        'deductions': float(round(deductions, 2)),
        'overtime_hours': float(overtime_hours),
        'overtime_amount': float(round(overtime_amount, 2)),
        'net_salary': float(round(net_salary, 2)),
        'payment_status': 'live',  # indicates real-time calculation
        'is_live': True,
    }


class StaffSalaryReportView(APIView):
    """Staff Salary Report — Real-time Monthly Sheet & Yearly Summary"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.staff.models import Staff
        
        report_type = request.query_params.get('type', 'monthly')
        month = int(request.query_params.get('month', date.today().month))
        year = int(request.query_params.get('year', date.today().year))
        
        active_staff = Staff.objects.filter(is_active=True).order_by('name')
        
        if report_type == 'monthly':
            return self._monthly_report(active_staff, month, year)
        else:
            return self._yearly_report(active_staff, year)
    
    def _monthly_report(self, staff_qs, month, year):
        month_name = calendar.month_name[month]
        days_in_month = calendar.monthrange(year, month)[1]
        
        staff_data = []
        totals = {'gross_salary': 0, 'deductions': 0, 'overtime_amount': 0, 'net_salary': 0}
        
        for s in staff_qs:
            row = _compute_staff_salary(s, month, year)
            staff_data.append(row)
            totals['gross_salary'] += row['gross_salary']
            totals['deductions'] += row['deductions']
            totals['overtime_amount'] += row['overtime_amount']
            totals['net_salary'] += row['net_salary']
        
        paid_count = sum(1 for s in staff_data if s['payment_status'] == 'paid')
        pending_count = sum(1 for s in staff_data if s['payment_status'] == 'pending')
        live_count = sum(1 for s in staff_data if s['payment_status'] == 'live')
        
        return Response({
            'month': month,
            'month_name': month_name,
            'year': year,
            'days_in_month': days_in_month,
            'staff_count': len(staff_data),
            'summary': {
                'total_gross': round(totals['gross_salary'], 2),
                'total_deductions': round(totals['deductions'], 2),
                'total_overtime': round(totals['overtime_amount'], 2),
                'total_net': round(totals['net_salary'], 2),
                'paid_count': paid_count,
                'pending_count': pending_count,
                'live_count': live_count,
            },
            'staff': staff_data,
        })
    
    def _yearly_report(self, staff_qs, year):
        staff_data = []
        month_totals = {m: 0 for m in range(1, 13)}
        grand_total = 0
        
        for s in staff_qs:
            months = {}
            yearly_total = 0
            
            for m in range(1, 13):
                row = _compute_staff_salary(s, m, year)
                net = row['net_salary']
                if net > 0:
                    months[m] = {
                        'net_salary': net,
                        'payment_status': row['payment_status'],
                    }
                    yearly_total += net
                    month_totals[m] += net
            
            staff_data.append({
                'id': s.id,
                'name': s.name,
                'role': s.get_role_display(),
                'months': months,
                'yearly_total': round(yearly_total, 2),
            })
            grand_total += yearly_total
        
        return Response({
            'year': year,
            'staff_count': len(staff_data),
            'grand_total': round(grand_total, 2),
            'month_totals': {calendar.month_abbr[m]: round(month_totals[m], 2) for m in range(1, 13)},
            'staff': staff_data,
        })


class StaffSalaryExcelView(APIView):
    """Export Staff Salary as Excel Sheet"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from apps.staff.models import Staff
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        report_type = request.query_params.get('type', 'monthly')
        month = int(request.query_params.get('month', date.today().month))
        year = int(request.query_params.get('year', date.today().year))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Styles
        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid')
        subheader_fill = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
        total_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        paid_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        pending_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        
        active_staff = Staff.objects.filter(is_active=True).order_by('name')
        
        if report_type == 'monthly':
            self._write_monthly(ws, active_staff, month, year, header_font, header_fill, subheader_fill, total_fill, paid_fill, pending_fill, thin_border)
            filename = f'Staff_Salary_{calendar.month_name[month]}_{year}.xlsx'
        else:
            self._write_yearly(ws, active_staff, year, header_font, header_fill, subheader_fill, total_fill, thin_border)
            filename = f'Staff_Salary_Yearly_{year}.xlsx'
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _write_monthly(self, ws, staff_qs, month, year, hf, hfill, sfill, tfill, pfill, pnfill, border):
        from openpyxl.styles import Font, Alignment, PatternFill
        
        live_fill = PatternFill(start_color='CFFAFE', end_color='CFFAFE', fill_type='solid')  # cyan for live
        month_name = calendar.month_name[month]
        ws.title = f'{month_name} {year}'
        
        # Title
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f'STAFF SALARY REPORT — {month_name.upper()} {year}'
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='6D28D9')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Headers
        headers = ['#', 'Staff Name', 'Role', 'Type', 'Working Days', 'Present', 'Absent', 'Half Days', 
                    'Sundays', 'Holidays', 'Gross (₹)', 'Deductions (₹)', 'Overtime (₹)', 'Net Salary (₹)', 'Status']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[row].height = 30
        
        # Data rows — uses real-time calculation
        totals = {'gross': 0, 'deductions': 0, 'overtime': 0, 'net': 0}
        for idx, s in enumerate(staff_qs, 1):
            row += 1
            sd = _compute_staff_salary(s, month, year)
            
            status_label = 'Paid' if sd['payment_status'] == 'paid' else 'Pending' if sd['payment_status'] == 'pending' else 'Live'
            
            data = [
                idx, sd['name'], sd['role'],
                'Daily' if sd['salary_type'] == 'daily' else 'Monthly',
                sd['total_working_days'],
                sd['days_present'],
                sd['days_absent'],
                sd['half_days'],
                sd['sundays'],
                sd['holidays_count'],
                sd['gross_salary'],
                sd['deductions'],
                sd['overtime_amount'],
                sd['net_salary'],
                status_label,
            ]
            
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if col in (11, 12, 13, 14) and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                # Color status column
                if col == 15:
                    if sd['payment_status'] == 'paid':
                        cell.fill = pfill
                    elif sd['payment_status'] == 'pending':
                        cell.fill = pnfill
                    else:
                        cell.fill = live_fill
            
            totals['gross'] += sd['gross_salary']
            totals['deductions'] += sd['deductions']
            totals['overtime'] += sd['overtime_amount']
            totals['net'] += sd['net_salary']
        
        # Totals row
        row += 1
        ws.cell(row=row, column=1, value='').fill = tfill
        total_cell = ws.cell(row=row, column=2, value='TOTAL')
        total_cell.font = Font(bold=True, size=12)
        total_cell.fill = tfill
        for col in range(3, 11):
            ws.cell(row=row, column=col).fill = tfill
        
        for col, key in [(11, 'gross'), (12, 'deductions'), (13, 'overtime'), (14, 'net')]:
            cell = ws.cell(row=row, column=col, value=totals[key])
            cell.font = Font(bold=True, size=11)
            cell.fill = tfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.number_format = '#,##0.00'
    
    def _write_yearly(self, ws, staff_qs, year, hf, hfill, sfill, tfill, border):
        from openpyxl.styles import Font, Alignment
        
        ws.title = f'Yearly {year}'
        
        # Title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f'STAFF YEARLY SALARY SUMMARY — {year}'
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='6D28D9')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Headers
        headers = ['#', 'Staff Name', 'Role'] + [calendar.month_abbr[m] for m in range(1, 13)] + ['Total (₹)']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[row].height = 30
        
        # Data rows — uses real-time calculation
        month_totals = {m: 0 for m in range(1, 13)}
        grand_total = 0
        
        for idx, s in enumerate(staff_qs, 1):
            row += 1
            yearly_total = 0
            
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=s.name).border = border
            ws.cell(row=row, column=3, value=s.get_role_display()).border = border
            
            for m in range(1, 13):
                sd = _compute_staff_salary(s, m, year)
                val = sd['net_salary']
                cell = ws.cell(row=row, column=3 + m, value=val if val > 0 else '-')
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if val > 0:
                    cell.number_format = '#,##0'
                yearly_total += val
                month_totals[m] += val
            
            total_cell = ws.cell(row=row, column=16, value=yearly_total)
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal='center')
            total_cell.border = border
            total_cell.number_format = '#,##0'
            grand_total += yearly_total
        
        # Totals row
        row += 1
        ws.cell(row=row, column=1).fill = tfill
        total_label = ws.cell(row=row, column=2, value='TOTAL')
        total_label.font = Font(bold=True, size=12)
        total_label.fill = tfill
        ws.cell(row=row, column=3).fill = tfill
        
        for m in range(1, 13):
            cell = ws.cell(row=row, column=3 + m, value=month_totals[m])
            cell.font = Font(bold=True)
            cell.fill = tfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            cell.number_format = '#,##0'
        
        grand_cell = ws.cell(row=row, column=16, value=grand_total)
        grand_cell.font = Font(bold=True, size=12, color='6D28D9')
        grand_cell.fill = tfill
        grand_cell.alignment = Alignment(horizontal='center')
        grand_cell.border = border
        grand_cell.number_format = '#,##0'

